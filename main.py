import tkinter as tk
import cv2
from PIL import ImageTk, Image
import numpy as np
import json
import face_recognition
import os
from openpyxl import load_workbook
from datetime import datetime, date

DEBUG = True
def debug(info):
    if DEBUG:
        print(info)
    else: pass


class Excel:
    def __init__(self, path):
        self.excel_file_path = path
        self.sample_sheet="sample"
        self.workbook = load_workbook(self.excel_file_path)
        self.sheet_for_today = date.today().strftime('%d-%m-%Y')

    def add_todays_sheet(self):
        self.workbook.copy_worksheet(self.workbook[self.sample_sheet])
        self.workbook[self.workbook.sheetnames[-1]].title = self.sheet_for_today
        self.workbook.save(self.excel_file_path)

    def fetch_info(self, id):
        id += 1
        if self.sheet_for_today in self.workbook.sheetnames:
            sheet = self.workbook[self.sheet_for_today]
            row_values = []
            for cell in sheet[id]:
                row_values.append(cell.value)
            return row_values
        else:
            self.add_todays_sheet()
            self.fetch_info(id)
    
    def add_entry(self, id, entry):
        if self.sheet_for_today in self.workbook.sheetnames:
            current_time = datetime.now().strftime('%H:%M:%S')
            sheet = self.workbook[self.sheet_for_today]
            sheet.cell(row=id+1,column=self.fetch_info(0).index(entry)+1).value = str(current_time)
            self.workbook.save(self.excel_file_path)
        else:
            self.add_todays_sheet()
            self.add_entry(id,entry)


class CameraApp:
    def __init__(self, window):
        self.window = window
        self.window.title("Attendance System")
        self.window.geometry("500x600")
       
        # Create a label to display the camera feed
        self.label = tk.Label(window)
        self.label.pack(padx=10, pady=10)

        # Create label variables to store the information labels
        self.name_label = tk.Label(window, text="Name: ", font=("Arial", 12, "bold"))
        self.name_label.pack(pady=5)
        self.reg_label = tk.Label(window, text="Registration number: ")
        self.reg_label.pack(pady=5)
        self.division_label = tk.Label(window, text="Division: ")
        self.division_label.pack(pady=5)
        self.roll_label = tk.Label(window, text="Roll Number: ")
        self.roll_label.pack(pady=5) 

        # Create a frame for the buttons
        self.buttons_frame = tk.Frame(window)
        self.buttons_frame.pack(padx=10, pady=10)

        # Create the "Start" button
        self.c_in = tk.Button(self.buttons_frame, text="Check in", command=self.check_in)
        self.c_in.pack(side=tk.LEFT, padx=30)

        # Create the "Stop" button
        self.c_out = tk.Button(self.buttons_frame, text="Check out", command=self.check_out)
        self.c_out.pack(side=tk.LEFT, padx=30)

        self.is_camera_running = False
        self.video_capture = None

        # Create an empty array to store the information
        self.information = ["id", "______________", "______________", "_", "__"]

        #FaceRecognition variables
        self.known_face_encodings=[]
        self.known_face_names=[]
        self.face_locations=[]
        self.face_encodings=[]
        self.face_names=[]
        self.process_this_frame = True

        #To get Encodings
        self.images_path = r"S:\Projects\Face Recognition Attendance System\faces"
        self.encodings_path = r"S:\Projects\Face Recognition Attendance System\face_encodings"

        #variables for json
        self.encoding_file_json = "encoding.json"
        self.json_file_data = ""

        self.excel_file = "database.xlsx"
        self.detected = None

        debug("[-] App initialized")
        #App initialization
        self.start_camera()
        self.update_ui()

    def write_json_encoding(self, data):
        debug("[-] Writing JSON file")
        with open(self.encoding_file_json, "w") as f:
            json.dump(data, f, indent=4)

    def read_json_encoding(self):
        debug("[-] Reading JSON file")
        with open(self.encoding_file_json, "r") as f:
            self.json_file_data = json.load(f)

    def train(self):
        debug("[-] Generating encodings for faces")
        json_data = {
            "id":[],
            "encoding_file_path":[]
        }
        for img_name in os.listdir(self.images_path):
            img_path = os.path.join(self.images_path, img_name)
            image = face_recognition.load_image_file(img_path)

            encoading_path = os.path.join(self.encodings_path, img_name.split(".")[0]+".npy")
            faceEncoding = face_recognition.face_encodings(image)[0]

            json_data["id"].append(img_name.split(".")[0])
            json_data["encoding_file_path"].append(encoading_path)

            np.save(encoading_path, faceEncoding)
        debug("[-] Encoadings generated")
        self.write_json_encoding(json_data)
  
    def get_known_faces(self):
        self.read_json_encoding()
        self.known_face_names = self.json_file_data["id"]
        for path in self.json_file_data["encoding_file_path"]:
            self.known_face_encodings.append(np.load(path))
    
    def start_camera(self):
        debug("[-] Starting camera")
        self.get_known_faces()
        if not self.is_camera_running:
            # Open the video capture from the camera
            self.video_capture = cv2.VideoCapture(0)
            self.is_camera_running = True

            # Start the video update loop
            debug("[-] Init camera update loop")
            self.update_camera()

    def stop_camera(self):
        if self.is_camera_running:
            # Release the video capture
            debug("[-] Stopping camera")
            self.video_capture.release()
            self.is_camera_running = False

    def check_in(self):
        Excel(self.excel_file).add_entry(self.detected, "Check In")
        debug(f"[*] {self.detected} --> Check In")
        # self.__init__(window)
        self.detected = None
        self.information = ["id", "______________", "______________", "_", "__"]
        self.face_names=[]
        # print("self.update_ui")
        self.update_ui()
        self.start_camera()

    def check_out(self):
        Excel(self.excel_file).add_entry(self.detected, "Check Out")
        debug(f"[*] {self.detected} --> Check Out")
        # self.__init__(window)
        self.detected = None
        self.information = ["id", "______________", "______________", "_", "__"]
        self.face_names=[]
        # print("self.update_ui")
        self.update_ui()
        self.start_camera()

    def update_camera(self):
        if self.is_camera_running:
            # Read a frame from the video capture
            ret, frame = self.video_capture.read()
            if ret:
                #FaceRecognition Code
                if self.process_this_frame:
                    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
                    rgb_small_frame = small_frame[:, :, ::-1]
                    self.face_locations = face_recognition.face_locations(rgb_small_frame)
                    self.face_encodings = face_recognition.face_encodings(rgb_small_frame, self.face_locations)

                    self.face_names = []
                    # print(self.face_encodings)
                    for face_encoding in self.face_encodings:
                        matches = face_recognition.compare_faces(self.known_face_encodings, face_encoding)
                        # print(matches)
                        name = "Unknown"
                        face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
                        best_match_index = np.argmin(face_distances)
                        # print(best_match_index)
                        if matches[best_match_index]:
                            name = self.known_face_names[best_match_index]
                            self.update_info(name)
                            debug(f"[@] Face detected : {name}")

                        self.face_names.append(name)

                self.process_this_frame = not self.process_this_frame
                for (top, right, bottom, left), name in zip(self.face_locations, self.face_names):
                    top *= 4
                    right *= 4
                    bottom *= 4
                    left *= 4

                    cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
                    cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
                    font = cv2.FONT_HERSHEY_DUPLEX
                    cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

                # Convert the OpenCV frame to PIL format
                image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                image = Image.fromarray(image)

                 # Resize the image to desired dimensions
                image = image.resize((500, 400))

                # Create an ImageTk object from the PIL image
                img_tk = ImageTk.PhotoImage(image)

                # Update the label with the new image
                self.label.imgtk = img_tk
                self.label.configure(image=img_tk)

                # print("All Good")
        # Repeat the update after a delay (in milliseconds)
        self.window.after(10, self.update_camera)

    def update_info(self, id):
        try:
            debug("[-] Updated Info")
            id = int(id)
            self.information = Excel(self.excel_file).fetch_info(id)
            self.detected = id
            self.update_ui()
            self.stop_camera()
        except Exception as e:
            debug(f"[#] Error: {e}")

    def update_ui(self):
        # Update the information labels with the array data
        debug("[-] Updating UI")
        self.name_label.config(text="Name: " + str(self.information[2]))
        self.reg_label.config(text="Registration number: " + str(self.information[1]))
        self.division_label.config(text="Division: " + str(self.information[3]))
        self.roll_label.config(text="Roll Number: " + str(self.information[4]))



if __name__=="__main__":
    # Create the Tkinter window
    window = tk.Tk()
    # Create the CameraApp instance
    app = CameraApp(window)

    # Uncomment following line to generate encodings 
    # app.train()
    
    # Run the Tkinter event loop    
    window.mainloop()
