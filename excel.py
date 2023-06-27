from openpyxl import load_workbook
from datetime import datetime, date

workbook = load_workbook("sheet1.xlsx")

def read(sheet_name, row_index):
    row_index += 1
    sheet = workbook[sheet_name]
    row_values = []
    for cell in sheet[row_index]:
        row_values.append(cell.value)
    return row_values

# def write_row_from_list(sheet_name, row_index, values):
#     row_index -=1
#     sheet = workbook[sheet_name]
#     for col_index, value in enumerate(values, start=1):
#         sheet.cell(row=row_index, column=col_index).value = value


# def write_data_to_cell(file_path, sample_sheet_name, column_title, row_id, data):
#     today = date.today()
#     today_sheet_name = today.strftime('%Y-%m-%d')
    
#     workbook = load_workbook(file_path)
#     sheet_names = workbook.sheetnames
    
#     # Check if today's sheet already exists
#     if today_sheet_name in sheet_names:
#         sheet = workbook[today_sheet_name]
#     else:
#         # Create a copy of the sample sheet if today's sheet doesn't exist
#         workbook.copy_worksheet(workbook[sample_sheet_name])
#         new_sheet_name = sheet_names[-1]  # Get the last sheet name (newly created sheet)
#         sheet = workbook[new_sheet_name]
#         sheet.title = today_sheet_name
    
#     # Find the column index based on the column title
#     column_index = None
#     for cell in sheet[1]:
#         if cell.value == column_title:
#             column_index = cell.column_letter
#             break
    
#     if column_index is None:
#         print(f"Column '{column_title}' not found.")
#         return
    
#     # Write data to the specified cell
#     cell_address = f"{column_index}{row_id}"
#     sheet[cell_address].value = data
    
#     workbook.save(file_path)

# data= read("27-06-2023", 1)
# print(data)

# # today = date.today()
# # today_time = today.(r'%Y-%m-%d')
# current_time = datetime.now().strftime('%H:%M:%S')

# sheet = workbook["27-06-2023"]
# sheet.cell(row=2, column=6).value = str(current_time)
# workbook.save("sheet1.xlsx")

# data[5] = "time"

# write_row_from_list("23-23-2003", 2, data)

# workbook.save("sheet1.xlsx")


def AddEntry(id, entry):
    sheetname = date.today().strftime('%d-%m-%Y')
    current_time = datetime.now().strftime('%H:%M:%S')

    sheet = workbook[sheetname]
    sheet.cell(row=id+1,column=read("sample", 0).index(entry)+1).value = str(current_time)
    workbook.save("sheet1.xlsx")

AddEntry(1, "Check Out")
